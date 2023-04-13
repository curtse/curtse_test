print('salees!')
a = 0
b = 4
c = (a+1.5) * b
print(c)
print('using Thonny')

x = 7.5
def square_x(x):
    x *= x
    print(x)

print(x)


# unrelated, excel sheets copy paste

import openpyxl

# Load Input1.xlsx workbook and Sheet1
wb_input1 = openpyxl.load_workbook('Input1.xlsx')
ws_input1 = wb_input1['Sheet1']

# Load Input2.xlsx workbook and Sheet2
wb_input2 = openpyxl.load_workbook('Input2.xlsx')
ws_input2 = wb_input2['Sheet2']

# Load Output.xlsx workbook and Output1 worksheet
wb_output1 = openpyxl.load_workbook('Output.xlsx')
ws_output1 = wb_output1['Output1']

# Load Output.xlsx workbook and Output2 worksheet
wb_output2 = openpyxl.load_workbook('Output.xlsx')
ws_output2 = wb_output2['Output2']

# Copy data from Sheet1 in Input1.xlsx to Output1.xlsx
for row in ws_input1.iter_rows(values_only=True):
    ws_output1.append(row)

# Copy data from Sheet2 in Input2.xlsx to Output2.xlsx
for row in ws_input2.iter_rows(values_only=True):
    ws_output2.append(row)

# Save changes to Output1.xlsx and Output2.xlsx
wb_output1.save('Output.xlsx')
wb_output2.save('Output.xlsx')

# Close all workbooks
wb_input1.close()
wb_input2.close()
wb_output1.close()
wb_output2.close()


### now the same using pandas instead of openpyxl

import pandas as pd

# Load Input1.xlsx workbook and Sheet1 into a DataFrame
df_input1 = pd.read_excel('Input1.xlsx', sheet_name='Sheet1')

# Load Input2.xlsx workbook and Sheet2 into a DataFrame
df_input2 = pd.read_excel('Input2.xlsx', sheet_name='Sheet2')

# Load Output.xlsx workbook and Output1 worksheet into a DataFrame
df_output1 = pd.read_excel('Output.xlsx', sheet_name='Output1')

# Load Output.xlsx workbook and Output2 worksheet into a DataFrame
df_output2 = pd.read_excel('Output.xlsx', sheet_name='Output2')

# Append data from Sheet1 in Input1.xlsx to Output1.xlsx
df_output1 = df_output1.append(df_input1, ignore_index=True)

# Append data from Sheet2 in Input2.xlsx to Output2.xlsx
df_output2 = df_output2.append(df_input2, ignore_index=True)

# Write changes to Output1.xlsx and Output2.xlsx
df_output1.to_excel('Output.xlsx', sheet_name='Output1', index=False)
df_output2.to_excel('Output.xlsx', sheet_name='Output2', index=False)


### get filename


import pandas as pd
import datetime

# Get yesterday's date
yesterday = datetime.date.today() - datetime.timedelta(days=1)

# Format yesterday's date as DDMMYY
date_str = yesterday.strftime('%d%m%y')

# Construct the filename with the formatted date
filename = f"sheet{date_str}.xlsx"

# Load the DataFrame from the Excel workbook
df = pd.read_excel(filename)

# Display
print(df)


### check file exists

import os

# Specify the filename to check
filename = "example.txt"

# Check if the file exists
if os.path.isfile(filename):
    print("found")
else:
    print("not found")

### Loop variant to check if all files exist. use os.path.isfile instead of glob 

import os
import glob

# Directory path where the Excel files are located
directory_path = 'path/to/directory'

# List of the eight specific Excel files to check for
files_to_check = ['file1.xlsx', 'file2.xlsx', 'file3.xlsx', 'file4.xlsx', 'file5.xlsx', 'file6.xlsx', 'file7.xlsx', 'file8.xlsx']

# Variable to keep track of the number of files found
files_found = 0

# Loop through the files_to_check list
for file in files_to_check:
    file_path = os.path.join(directory_path, file)  # Get the full file path by joining directory path and file name
    if glob.glob(file_path):  # Check if the file exists
        files_found += 1
    else:
        print("Fail")
        break

# Check if all files were found
if files_found == len(files_to_check):
    print("Number of files found: ", files_found)



